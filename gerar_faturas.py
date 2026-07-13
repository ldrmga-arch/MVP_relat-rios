#!/usr/bin/env python3
"""
gerar_faturas.py — MDN Porto Rico
Gera faturas no formato do Modelo de exemplo (aba PR RR QD 23 LT 14).

Tabela esquerda (col C, linhas 19-27):
  Linhas 19-21  SERVIÇOS MDN      A / B / C
  Linhas 22-24  REPOSIÇÃO DE ITEM A / B / C
  Linhas 25-27  PRODUTOS LIMPEZA  A / B / C

Tabela direita (CI):
  Col F  cota (A/B/C)
  Col H  dia 15 do mês anterior à emissão
  Col I  consumo energia (kWh, leitura atual - anterior)
  Col J  "1.4.12 - Energia - CI"
  Col K  período (dd/mm/aaaa à dd/mm/aaaa)

Uso:
    pip install openpyxl
    python gerar_faturas.py
"""

import json, re, sys
from collections import defaultdict
from datetime import date, timedelta
import urllib.request, urllib.parse
import threading

import openpyxl
from openpyxl import load_workbook

# ─── CONFIGURAÇÃO ─────────────────────────────────────────────────────────────
CLIENT_ID     = "gmntttOuval2d0etI6bB0Qq5dbdB-VYXjZcQygyNF4w"
CLIENT_SECRET = "XJPw0LVWIuouVXFBn6KmE9CCC0C8ooXMz1GgVPLJO-0"
TOKEN_URL     = "https://app.pipefy.com/oauth/token"
GRAPHQL_URL   = "https://api.pipefy.com/graphql"

FASE_MANUT_OK   = "329509955"
FASE_ZEL_OK     = "329593953"   # Zeladoria — Concluído
FASE_ENERGIA_OK = "340612497"
DB_COTISTAS     = "304903871"

TEMPLATE_FILE  = "Modelo de exemplo.xlsx"
TEMPLATE_SHEET = "PR RR QD 23 LT 14"

# Imóvel de teste — None = todos
APENAS_IMOVEL = "PR RR Q02 L14"

# Período manutenção / zeladoria
PERIODO_MANUAL_INI = "2026-05-27"
PERIODO_MANUAL_FIM = "2026-06-26"

# Datas da fatura
DATA_EMISSAO    = date(2026, 7, 1)
DATA_VENCIMENTO = date(2026, 7, 15)

# Cotistas ignorados e alias (espelha relatorio_cotistas.py)
COTISTAS_IGNORAR: set = {
    "Rafael França",
    "Marcel Sanches",
    "Gustavo Pagani",
    "Maria Wilma Wohlers da Silva",
}
COTISTAS_ALIAS: dict = {
    "CARINA MUZULAN": "Carina / Clovis Muzulan",
    "Clovis Muzulan": "Carina / Clovis Muzulan",
}

# Tipos de zeladoria (fase Concluído) → categoria na fatura
TIPO_ZEL_CATEGORIA = {
    "Reposição de itens":  "REPOSIÇÃO DE ITEM",
    "Limpeza Pet":         "PRODUTOS LIMPEZA",
    "Produtos de limpeza": "PRODUTOS LIMPEZA",
}

# ─── PERÍODO ──────────────────────────────────────────────────────────────────
PERIODO_INI = date.fromisoformat(PERIODO_MANUAL_INI)
PERIODO_FIM = date.fromisoformat(PERIODO_MANUAL_FIM)

def _ultima_quarta_ate_dia10(year, month):
    d = date(year, month, 10)
    return d - timedelta(days=(d.weekday() - 2) % 7)

_bm, _by = PERIODO_FIM.month, PERIODO_FIM.year
_pm = _bm - 1 if _bm > 1 else 12
_py = _by  if _bm > 1 else _by - 1
ENERGIA_INI = _ultima_quarta_ate_dia10(_py, _pm)
ENERGIA_FIM = _ultima_quarta_ate_dia10(_by, _bm)
ENERGIA_PERIODO_STR = (
    f"{ENERGIA_INI.strftime('%d/%m/%Y')} à {ENERGIA_FIM.strftime('%d/%m/%Y')}"
)

# Dia 15 do mês anterior à emissão (data de referência energia)
_rm = DATA_EMISSAO.month - 1 if DATA_EMISSAO.month > 1 else 12
_ry = DATA_EMISSAO.year  if DATA_EMISSAO.month > 1 else DATA_EMISSAO.year - 1
DATA_REF_ENERGIA = date(_ry, _rm, 15)

# ─── API ──────────────────────────────────────────────────────────────────────
def _get_token():
    payload = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID, "client_secret": CLIENT_SECRET,
    }).encode()
    req = urllib.request.Request(TOKEN_URL, data=payload, method="POST")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read())["access_token"]

TOKEN = _get_token()
print("✓ Token OK")

def gql(query, variables=None):
    body = json.dumps({"query": query, "variables": variables or {}}).encode()
    req = urllib.request.Request(GRAPHQL_URL, data=body, method="POST")
    req.add_header("Authorization", f"Bearer {TOKEN}")
    req.add_header("Content-Type", "application/json")
    with urllib.request.urlopen(req, timeout=90) as r:
        d = json.loads(r.read())
    if "data" not in d:
        raise RuntimeError(json.dumps(d))
    return d["data"]

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def normalizar(nome: str) -> str:
    nome = (nome or "").strip()
    return COTISTAS_ALIAS.get(nome, nome)

def parse_json_array(val) -> str:
    if not val:
        return ""
    try:
        lst = json.loads(val)
        return lst[0] if lst else ""
    except Exception:
        return str(val).strip('[]"\'')

def parse_float(val) -> float:
    if not val:
        return 0.0
    nums = re.findall(r'\d+(?:[.,]\d+)?', str(val).replace(",", "."))
    return float(nums[-1].replace(",", ".")) if nums else 0.0

def to_date(s):
    if not s:
        return None
    try:
        return date.fromisoformat(str(s)[:10])
    except Exception:
        return None

def parse_energy(obs, field):
    for v in (field, obs):
        if v:
            nums = re.findall(r'\d+(?:[.,]\d+)?', str(v).strip())
            if nums:
                return parse_float(nums[-1])
    return None

def in_period(d):
    return d is not None and PERIODO_INI <= d <= PERIODO_FIM

def in_energia_period(d):
    return d is not None and ENERGIA_INI < d <= ENERGIA_FIM

def card_fields(card):
    return {f["name"]: (f.get("value") or "") for f in card.get("fields", [])}

# ─── QUERIES ──────────────────────────────────────────────────────────────────
_Q_PHASE = """
query($pid: ID!, $n: Int!, $after: String, $from: String, $to: String) {
  phase(id: $pid) {
    cards(first: $n, after: $after, filter: {created_at: {from: $from, to: $to}}) {
      pageInfo { hasNextPage endCursor }
      edges { node { id title created_at fields { name value } } }
    }
  }
}"""

_Q_DB = """
query($tid: ID!, $n: Int!, $after: String) {
  table_records(table_id: $tid, first: $n, after: $after) {
    pageInfo { hasNextPage endCursor }
    edges { node { title record_fields { name value array_value } } }
  }
}"""

def fetch_phase(phase_id, date_from=None, date_to=None):
    cards, after = [], None
    _from = date_from.isoformat() if date_from else None
    _to   = date_to.isoformat()   if date_to   else None
    while True:
        try:
            d = gql(_Q_PHASE, {"pid": phase_id, "n": 50, "after": after,
                                "from": _from, "to": _to})
            page = d["phase"]["cards"]
        except RuntimeError:
            # fallback sem filtro de data
            _Q2 = """
query($pid: ID!, $n: Int!, $after: String) {
  phase(id: $pid) {
    cards(first: $n, after: $after) {
      pageInfo { hasNextPage endCursor }
      edges { node { id title created_at fields { name value } } }
    }
  }
}"""
            d = gql(_Q2, {"pid": phase_id, "n": 50, "after": after})
            page = d["phase"]["cards"]
        for e in page["edges"]:
            cards.append(e["node"])
        if not page["pageInfo"]["hasNextPage"]:
            break
        after = page["pageInfo"]["endCursor"]
    return cards

def fetch_db(tid):
    records, after = [], None
    while True:
        d = gql(_Q_DB, {"tid": tid, "n": 50, "after": after})
        page = d["table_records"]
        for e in page["edges"]:
            records.append(e["node"])
        if not page["pageInfo"]["hasNextPage"]:
            break
        after = page["pageInfo"]["endCursor"]
    return records

# ─── CARREGAR DADOS ───────────────────────────────────────────────────────────
print(f"Período: {PERIODO_INI:%d/%m/%Y} — {PERIODO_FIM:%d/%m/%Y}")
print(f"Energia: {ENERGIA_INI:%d/%m/%Y} — {ENERGIA_FIM:%d/%m/%Y}")
print(f"Data de referência energia: {DATA_REF_ENERGIA:%d/%m/%Y}")
print("\n── Carregando (paralelo) ──────────────────────────────────")

_res = {}
def _load(key, fn, *args, **kw):
    _res[key] = fn(*args, **kw)

_energia_from = ENERGIA_INI - timedelta(days=45)
threads = [
    threading.Thread(target=_load, args=("cot",   fetch_db,    DB_COTISTAS)),
    threading.Thread(target=_load, args=("manut",  fetch_phase, FASE_MANUT_OK),
                     kwargs={"date_from": PERIODO_INI, "date_to": PERIODO_FIM}),
    threading.Thread(target=_load, args=("zel",    fetch_phase, FASE_ZEL_OK),
                     kwargs={"date_from": PERIODO_INI, "date_to": PERIODO_FIM}),
    threading.Thread(target=_load, args=("energ",  fetch_phase, FASE_ENERGIA_OK),
                     kwargs={"date_from": _energia_from, "date_to": PERIODO_FIM}),
]
for t in threads: t.start()
for t in threads: t.join()

cotistas_raw = _res["cot"]
manut_all    = _res["manut"]
zel_all      = _res["zel"]
energ_all    = _res["energ"]
print(f"  Cotistas: {len(cotistas_raw)} | Manutenção: {len(manut_all)} | "
      f"Zeladoria: {len(zel_all)} | Energia: {len(energ_all)}")

# ─── MAPEAMENTO COTISTAS ──────────────────────────────────────────────────────
# cotas_imovel[imovel] = {cota_letter: cotista_name}
cotas_imovel: dict[str, dict[str, str]] = defaultdict(dict)
# cotista_para_origem[cotista] = imovel (home)
cotista_para_origem: dict[str, str] = {}

for rec in cotistas_raw:
    nome_raw = (rec.get("title") or "").strip()
    if not nome_raw or nome_raw in COTISTAS_IGNORAR:
        continue
    nome = normalizar(nome_raw)
    fmap = {f["name"]: f for f in rec.get("record_fields", [])}

    # Imóvel: usar 'value' (título), não array_value (IDs)
    f_im = fmap.get("Imóvel")
    if not f_im:
        continue
    v = (f_im.get("value") or "").strip()
    try:
        im_list = json.loads(v) if v else []
        if not isinstance(im_list, list):
            im_list = [im_list]
    except Exception:
        im_list = [v] if v else []

    # Cota
    f_cota = fmap.get("Cota")
    cota = ((f_cota.get("value") or "").strip() if f_cota else "").upper()

    for imovel in im_list:
        imovel = str(imovel).strip()
        if not imovel:
            continue
        if cota:
            cotas_imovel[imovel][cota] = nome
        if nome not in cotista_para_origem:
            cotista_para_origem[nome] = imovel

print(f"  Imóveis mapeados: {len(cotas_imovel)}")

# ─── PROCESSAR CARDS ──────────────────────────────────────────────────────────

# Manutenção
manut_periodo: list[dict] = []
for card in manut_all:
    dt = to_date(card["created_at"])
    if not in_period(dt):
        continue
    f = card_fields(card)
    custeio = f.get("Responsável pelo custeio", "")
    if custeio not in ("Rateio entre cotistas", "Cotista"):
        continue
    imovel = parse_json_array(f.get("Imóveis", ""))
    if APENAS_IMOVEL and imovel != APENAS_IMOVEL:
        continue
    manut_periodo.append({
        "imovel":      imovel,
        "valor":       parse_float(f.get("Valor total da manutenção", "0")),
        "custeio":     custeio,
        "cotista_esp": normalizar(parse_json_array(f.get("Selecione o cliente", ""))),
    })

# Zeladoria
zel_periodo: list[dict] = []
for card in zel_all:
    dt = to_date(card["created_at"])
    if not in_period(dt):
        continue
    f = card_fields(card)
    tipo = f.get("Seleção de lista", "")
    if tipo not in TIPO_ZEL_CATEGORIA:
        continue
    # "Cobrança para" é usado na fase Administrativo (Produtos de limpeza)
    # "Responsável pela cobrança" é usado nas demais fases
    cobranca = f.get("Responsável pela cobrança", "") or f.get("Cobrança para", "")
    if cobranca not in ("Rateio entre cotistas", "Cotista"):
        continue
    imovel = parse_json_array(f.get("Selecione o imóvel", ""))
    if APENAS_IMOVEL and imovel != APENAS_IMOVEL:
        continue
    zel_periodo.append({
        "imovel":      imovel,
        "tipo":        tipo,
        "categoria":   TIPO_ZEL_CATEGORIA[tipo],
        "valor":       parse_float(f.get("Valor da cobrança", "0")),
        "custeio":     cobranca,
        "cotista_esp": normalizar(parse_json_array(f.get("Selecione o cotista:", ""))),
    })

# Energia: leituras por imóvel de uso
_leituras: dict[str, list] = defaultdict(list)
for card in energ_all:
    f = card_fields(card)
    imovel_uso = parse_json_array(f.get("Selecione o imóvel", ""))
    leitura    = parse_energy(f.get("Observação:", ""),
                              f.get("Digite o valor no relógio de energia:", ""))
    cotista    = normalizar(parse_json_array(f.get("Cotista", "")))
    _leituras[imovel_uso].append({
        "data":    to_date(card["created_at"]),
        "leitura": leitura,
        "cotista": cotista,
    })

# Calcular consumo por período, atribuindo ao imóvel de ORIGEM do cotista
energia_periodo: list[dict] = []
for imovel_uso, cards in _leituras.items():
    cards.sort(key=lambda x: x["data"] or date.min)
    for i, card in enumerate(cards):
        if not in_energia_period(card["data"]):
            continue
        prev    = cards[i - 1] if i > 0 else None
        leit_ini = prev["leitura"] if prev else None
        consumo  = (
            card["leitura"] - leit_ini
            if card["leitura"] is not None and leit_ini is not None
            else None
        )
        cotista       = card["cotista"]
        imovel_origem = cotista_para_origem.get(cotista, "")
        data_ini      = prev["data"] if prev else None
        periodo_uso   = (
            f"{data_ini.strftime('%d/%m')} à {card['data'].strftime('%d/%m')}"
            if data_ini and card["data"] else ""
        )
        energia_periodo.append({
            "cotista":       cotista,
            "imovel_uso":    imovel_uso,
            "imovel_origem": imovel_origem,
            "consumo":       consumo,
            "periodo":       periodo_uso,
        })

print(f"  Manutenção: {len(manut_periodo)} | Zeladoria: {len(zel_periodo)} "
      f"| Energia: {len(energia_periodo)}")

# ─── CALCULA VALORES POR IMÓVEL ───────────────────────────────────────────────
def calcular_imovel(imovel: str):
    cotas = cotas_imovel.get(imovel, {})  # {cota_letter: cotista_name}
    n     = max(len(cotas), 1)
    # Mapa reverso para este imóvel: cotista → cota
    cot2cota = {v: k for k, v in cotas.items()}

    # ── SERVIÇOS MDN (Manutenção) ───────────────────────────────────────────
    m_all    = [r for r in manut_periodo if r["imovel"] == imovel]
    m_rateio = sum(r["valor"] for r in m_all if r["custeio"] == "Rateio entre cotistas")
    m_ind    = [r for r in m_all if r["custeio"] == "Cotista"]

    serv_mdn: dict[str, float] = {}
    for cota, cotista in cotas.items():
        ind = sum(r["valor"] for r in m_ind if r["cotista_esp"] == cotista)
        serv_mdn[cota] = round(m_rateio / n + ind, 2)

    # ── REPOSIÇÃO / PRODUTOS (Zeladoria) ────────────────────────────────────
    repos_item:    dict[str, float] = {}
    prod_limpeza:  dict[str, float] = {}

    for categoria, store in [("REPOSIÇÃO DE ITEM", repos_item),
                              ("PRODUTOS LIMPEZA",  prod_limpeza)]:
        z_all  = [r for r in zel_periodo
                  if r["imovel"] == imovel and r["categoria"] == categoria]
        z_rat  = sum(r["valor"] for r in z_all
                     if r["custeio"] == "Rateio entre cotistas")
        for cota, cotista in cotas.items():
            z_ind = sum(r["valor"] for r in z_all
                        if r["custeio"] == "Cotista"
                        and r["cotista_esp"] == cotista)
            store[cota] = round(z_rat / n + z_ind, 2)

    # ── ENERGIA ─────────────────────────────────────────────────────────────
    # Agrupa consumo por cota (pode ter várias leituras)
    e_por_cota: dict[str, float] = defaultdict(float)
    e_periodos:  dict[str, list]  = defaultdict(list)
    for r in energia_periodo:
        if r["imovel_origem"] != imovel:
            continue
        cotista = r["cotista"]
        cota    = cot2cota.get(cotista, "")
        if cota and r["consumo"] is not None:
            e_por_cota[cota] += r["consumo"]
            e_periodos[cota].append(r["periodo"])

    return cotas, n, serv_mdn, repos_item, prod_limpeza, dict(e_por_cota), dict(e_periodos)

# ─── GERAR EXCEL ──────────────────────────────────────────────────────────────
print("\n── Gerando Excel ──────────────────────────────────────────")
wb = load_workbook(TEMPLATE_FILE)
ws_tpl = wb[TEMPLATE_SHEET]

# Imóveis a processar
if APENAS_IMOVEL:
    imoveis = [APENAS_IMOVEL]
else:
    imoveis = sorted(set(
        [r["imovel"]        for r in manut_periodo] +
        [r["imovel"]        for r in zel_periodo]   +
        [r["imovel_origem"] for r in energia_periodo if r["imovel_origem"]]
    ))

def _safe_title(name: str) -> str:
    for ch in "[]:*?/\\":
        name = name.replace(ch, "")
    return name[:31].strip()

for imovel in imoveis:
    cotas, n, serv_mdn, repos_item, prod_limpeza, e_por_cota, e_periodos = \
        calcular_imovel(imovel)

    # Copia aba-template
    ws = wb.copy_worksheet(ws_tpl)
    ws.title = _safe_title(imovel)

    sorted_cotas = sorted(cotas.keys())   # ["A", "B", "C"]

    # ── Cabeçalho ─────────────────────────────────────────────────────────
    ws["A8"] = f"MONDONEX - {imovel}"
    ws["B8"] = n
    ws["D5"] = DATA_EMISSAO
    ws["D8"] = DATA_VENCIMENTO

    # ── Tabela esquerda — linhas variáveis (19-27) ─────────────────────────
    itens = [
        (serv_mdn,   19),   # SERVIÇOS MDN:      linhas 19, 20, 21
        (repos_item, 22),   # REPOSIÇÃO DE ITEM: linhas 22, 23, 24
        (prod_limpeza, 25), # PRODUTOS LIMPEZA:  linhas 25, 26, 27
    ]
    for data_dict, base_row in itens:
        for i, cota in enumerate(sorted_cotas):
            r = base_row + i
            ws.cell(r, 2).value = cota                      # col B: letra cota
            ws.cell(r, 3).value = data_dict.get(cota, 0)   # col C: valor R$

    # Limpar linhas de cota que excederem o número de cotistas
    for data_dict, base_row in itens:
        for i in range(len(sorted_cotas), 3):
            ws.cell(base_row + i, 2).value = None
            ws.cell(base_row + i, 3).value = None

    # ── Tabela direita — energia (CI) ──────────────────────────────────────
    # Limpa dados de exemplo (linhas 2-29, colunas F=6 a L=12)
    for r in range(2, 30):
        for c in range(6, 13):
            ws.cell(r, c).value = None

    # Preenche uma linha por cota com consumo > 0
    r_ci = 2
    for cota in sorted_cotas:
        consumo = e_por_cota.get(cota, 0.0)
        if consumo <= 0:
            continue
        # Período: junta todos os subperíodos desta cota
        periodos = e_periodos.get(cota, [])
        if periodos:
            periodo_str = periodos[0] if len(periodos) == 1 else \
                          f"{periodos[0].split(' à ')[0]} à {periodos[-1].split(' à ')[-1]}"
        else:
            periodo_str = ENERGIA_PERIODO_STR

        ws.cell(r_ci, 6).value  = cota                       # F: cota
        ws.cell(r_ci, 8).value  = DATA_REF_ENERGIA           # H: data ref
        ws.cell(r_ci, 9).value  = round(consumo, 2)          # I: kWh
        ws.cell(r_ci, 10).value = "1.4.12 - Energia - CI"    # J: conta
        ws.cell(r_ci, 11).value = ENERGIA_PERIODO_STR        # K: período
        r_ci += 1

    print(f"  ✓ {imovel}  ({n} cotas | MDN R${sum(serv_mdn.values()):.2f} "
          f"| Rep R${sum(repos_item.values()):.2f} "
          f"| Enrg {sum(e_por_cota.values()):.1f} kWh)")

# Remove abas de exemplo (mantém Plano de conta e TOTALIZADOR)
for sheet_name in ["PR RR QD02 LT14", TEMPLATE_SHEET]:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

filename = f"Faturas_{PERIODO_INI:%d%m}-{PERIODO_FIM:%d%m%Y}.xlsx"
wb.save(filename)
print(f"\n✅  Arquivo gerado: {filename}")
