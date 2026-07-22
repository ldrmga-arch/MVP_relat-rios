#!/usr/bin/env python3
"""
Relatórios API — MDN Porto Rico
GET /relatorios/api/gerar?ini=2026-05-27&fim=2026-06-26[&imovel=PR+RR+Q02+L14]
Gera o relatório mensal Excel e retorna como download.
"""
import json, os, re, io, threading
from collections import defaultdict
from datetime import date, timedelta
import urllib.request, urllib.parse
from flask import Flask, jsonify, request, Response
from dotenv import load_dotenv

load_dotenv()

CLIENT_ID     = os.environ.get("PIPEFY_CLIENT_ID", "") or os.environ.get("CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("PIPEFY_CLIENT_SECRET", "") or os.environ.get("CLIENT_SECRET", "")
TOKEN_URL     = "https://app.pipefy.com/oauth/token"
GRAPHQL_URL   = "https://api.pipefy.com/graphql"

FASE_MANUT_OK    = "329509955"
FASE_ZEL_OK      = "329593953"
FASE_ENERGIA_OK  = "340612497"
DB_IMOVEIS       = "304884712"
DB_COTISTAS      = "304903871"
TIPOS_ZELADORIA  = {"Reposição de itens", "Limpeza Pet"}

COTISTAS_IGNORAR = {
    "Rafael França", "Marcel Sanches",
    "Gustavo Pagani", "Maria Wilma Wohlers da Silva",
}
COTISTAS_ALIAS = {
    "CARINA MUZULAN": "Carina / Clovis Muzulan",
    "Clovis Muzulan": "Carina / Clovis Muzulan",
}

_Q_PHASE = """
query($pid: ID!, $n: Int!, $after: String, $from: String, $to: String) {
  phase(id: $pid) {
    cards(first: $n, after: $after, filter: {created_at: {from: $from, to: $to}}) {
      pageInfo { hasNextPage endCursor }
      edges { node { id title created_at fields { name value } } }
    }
  }
}"""

_Q_PHASE_NF = """
query($pid: ID!, $n: Int!, $after: String) {
  phase(id: $pid) {
    cards(first: $n, after: $after) {
      pageInfo { hasNextPage endCursor }
      edges { node { id title created_at fields { name value } } }
    }
  }
}"""

_Q_DB = """
query($tid: ID!, $n: Int!, $after: String) {
  table_records(table_id: $tid, first: $n, after: $after) {
    pageInfo { hasNextPage endCursor }
    edges { node { id title record_fields { name value array_value } } }
  }
}"""

app = Flask(__name__)

# ─── Pipefy helpers ───────────────────────────────────────────────────────────

def _get_token():
    data = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID, "client_secret": CLIENT_SECRET,
    }).encode()
    req = urllib.request.Request(TOKEN_URL, data=data, method="POST")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read())["access_token"]


def _gql(token, query, variables=None):
    body = json.dumps({"query": query, "variables": variables or {}}).encode()
    req = urllib.request.Request(GRAPHQL_URL, data=body, method="POST")
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Content-Type", "application/json")
    with urllib.request.urlopen(req, timeout=90) as r:
        data = json.loads(r.read())
    if "errors" in data:
        raise RuntimeError(f"GraphQL error: {data['errors']}")
    return data["data"]


def _fetch_phase(token, phase_id, date_from=None, date_to=None):
    cards, after = [], None
    use_filter = bool(date_from and date_to)
    while True:
        try:
            if use_filter:
                d = _gql(token, _Q_PHASE, {
                    "pid": phase_id, "n": 50, "after": after,
                    "from": date_from, "to": date_to,
                })
            else:
                d = _gql(token, _Q_PHASE_NF, {"pid": phase_id, "n": 50, "after": after})
        except RuntimeError:
            if use_filter:
                use_filter, after, cards = False, None, []
                continue
            raise
        page = d["phase"]["cards"]
        for e in page["edges"]:
            cards.append(e["node"])
        if not page["pageInfo"]["hasNextPage"]:
            break
        after = page["pageInfo"]["endCursor"]
    return cards


def _fetch_db(token, table_id):
    records, after = [], None
    while True:
        d = _gql(token, _Q_DB, {"tid": table_id, "n": 50, "after": after})
        page = d["table_records"]
        for e in page["edges"]:
            records.append(e["node"])
        if not page["pageInfo"]["hasNextPage"]:
            break
        after = page["pageInfo"]["endCursor"]
    return records

# ─── Data helpers ──────────────────────────────────────────────────────────────

def _normalizar(nome):
    nome = (nome or "").strip()
    return COTISTAS_ALIAS.get(nome, nome)


def _parse_json_array(val):
    if not val:
        return ""
    try:
        lst = json.loads(val)
        return lst[0] if lst else ""
    except Exception:
        return str(val).strip('[]"\'')


def _parse_float(val):
    if not val:
        return 0.0
    try:
        return float(str(val).replace(",", ".").strip())
    except Exception:
        return 0.0


def _parse_energy(obs_val, field_val):
    for v in (field_val, obs_val):
        if v:
            nums = re.findall(r'\d+(?:[.,]\d+)?', str(v).strip())
            if nums:
                return _parse_float(nums[-1])
    return None


def _to_date(s):
    if not s:
        return None
    try:
        return date.fromisoformat(str(s)[:10])
    except Exception:
        return None


def _ultima_quarta_ate_dia10(year, month):
    d = date(year, month, 10)
    return d - timedelta(days=(d.weekday() - 2) % 7)


def _safe_sheet(name):
    for ch in "[]:*?/\\":
        name = name.replace(ch, "")
    return name[:31].strip()

# ─── Excel generation ─────────────────────────────────────────────────────────

def _gerar_relatorio(periodo_ini, periodo_fim, apenas_imovel=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Energy period
    _bm, _by = periodo_fim.month, periodo_fim.year
    _pm = _bm - 1 if _bm > 1 else 12
    _py = _by if _bm > 1 else _by - 1
    energia_ini = _ultima_quarta_ate_dia10(_py, _pm)
    energia_fim = _ultima_quarta_ate_dia10(_by, _bm)

    # Fetch all data in parallel
    token = _get_token()
    _res = {}

    def _load(key, fn, *args, **kw):
        _res[key] = fn(token, *args, **kw)

    _energia_from = energia_ini - timedelta(days=45)
    threads = [
        threading.Thread(target=_load, args=("imoveis",  _fetch_db,    DB_IMOVEIS)),
        threading.Thread(target=_load, args=("cotistas", _fetch_db,    DB_COTISTAS)),
        threading.Thread(target=_load, args=("manut",    _fetch_phase, FASE_MANUT_OK),
                         kwargs={"date_from": periodo_ini.isoformat(), "date_to": periodo_fim.isoformat()}),
        threading.Thread(target=_load, args=("zel",      _fetch_phase, FASE_ZEL_OK),
                         kwargs={"date_from": periodo_ini.isoformat(), "date_to": periodo_fim.isoformat()}),
        threading.Thread(target=_load, args=("energ",    _fetch_phase, FASE_ENERGIA_OK),
                         kwargs={"date_from": _energia_from.isoformat(), "date_to": periodo_fim.isoformat()}),
    ]
    for t in threads: t.start()
    for t in threads: t.join()

    imoveis_raw  = _res["imoveis"]
    cotistas_raw = _res["cotistas"]
    manut_all    = _res["manut"]
    zel_all      = _res["zel"]
    energia_all  = _res["energ"]

    # Build cotistas → imovel mapping
    imoveis_cotistas: dict = {}
    for rec in cotistas_raw:
        nome_raw = (rec.get("title") or "").strip()
        if not nome_raw or nome_raw in COTISTAS_IGNORAR:
            continue
        nome = _normalizar(nome_raw)
        fmap = {f["name"]: f for f in rec.get("record_fields", [])}
        f_im = fmap.get("Imóvel")
        if not f_im:
            continue
        v = (f_im.get("value") or "").strip()
        try:
            parsed = json.loads(v)
            imoveis_list = [str(x).strip() for x in (parsed if isinstance(parsed, list) else [parsed])]
        except Exception:
            imoveis_list = [v] if v else []
        for imovel in imoveis_list:
            imovel = str(imovel).strip()
            if not imovel:
                continue
            if imovel not in imoveis_cotistas:
                imoveis_cotistas[imovel] = []
            if nome not in imoveis_cotistas[imovel]:
                imoveis_cotistas[imovel].append(nome)

    cotista_para_origem: dict = {}
    for im, cots in imoveis_cotistas.items():
        for cot in cots:
            if cot not in cotista_para_origem:
                cotista_para_origem[cot] = im

    def in_period(d):
        return d is not None and periodo_ini <= d <= periodo_fim

    def in_energia_period(d):
        return d is not None and energia_ini < d <= energia_fim

    def card_fields(card):
        return {f["name"]: (f.get("value") or "") for f in card.get("fields", [])}

    # Process manutenção
    manut_periodo = []
    for card in manut_all:
        dt = _to_date(card["created_at"])
        if not in_period(dt):
            continue
        f = card_fields(card)
        custeio = f.get("Responsável pelo custeio", "")
        if custeio not in ("Rateio entre cotistas", "Cotista"):
            continue
        manut_periodo.append({
            "id":          card["id"],
            "data":        dt,
            "imovel":      _parse_json_array(f.get("Imóveis", "")),
            "descricao":   (f.get("Descreva a solicitação") or f.get("Descreva o problema:", ""))[:200],
            "solucao":     (f.get("Descreva a solução")     or f.get("Descreva a solução:", ""))[:200],
            "valor":       _parse_float(f.get("Valor total da manutenção", "0")),
            "custeio":     custeio,
            "cotista_esp": _normalizar(_parse_json_array(f.get("Selecione o cliente", ""))),
        })

    # Process zeladoria
    zel_periodo = []
    for card in zel_all:
        dt = _to_date(card["created_at"])
        if not in_period(dt):
            continue
        f = card_fields(card)
        if f.get("Seleção de lista", "") not in TIPOS_ZELADORIA:
            continue
        cobranca = f.get("Responsável pela cobrança", "")
        if cobranca not in ("Rateio entre cotistas", "Cotista"):
            continue
        zel_periodo.append({
            "id":          card["id"],
            "data":        dt,
            "imovel":      _parse_json_array(f.get("Selecione o imóvel", "")),
            "tipo":        f.get("Seleção de lista", ""),
            "descricao":   f.get("Descreva a solicitação:", "")[:200],
            "valor":       _parse_float(f.get("Valor da cobrança", "0")),
            "custeio":     cobranca,
            "cotista_esp": _normalizar(_parse_json_array(f.get("Selecione o cotista:", ""))),
        })

    # Process energia
    _ep_imovel: dict = defaultdict(list)
    for card in energia_all:
        f = card_fields(card)
        iu = _parse_json_array(f.get("Selecione o imóvel", ""))
        if not iu:
            continue
        _ep_imovel[iu].append({
            "id":      card["id"],
            "data":    _to_date(card["created_at"]),
            "leitura": _parse_energy(f.get("Observação:", ""),
                                     f.get("Digite o valor no relógio de energia:", "")),
            "cotista": _normalizar(_parse_json_array(f.get("Cotista", ""))),
        })

    for im in _ep_imovel:
        _ep_imovel[im].sort(key=lambda x: x["data"] or date.min)

    energia_periodo = []
    for imovel_uso, cards in _ep_imovel.items():
        for i, card in enumerate(cards):
            if not in_energia_period(card["data"]):
                continue
            prev        = cards[i - 1] if i > 0 else None
            leitura_ini = prev["leitura"] if prev else None
            data_ini    = prev["data"]    if prev else None
            consumo = (card["leitura"] - leitura_ini
                       if card["leitura"] is not None and leitura_ini is not None
                       else None)
            periodo_uso = (f"{data_ini.strftime('%d/%m')} a {card['data'].strftime('%d/%m')}"
                           if data_ini and card["data"] else "")
            cotista      = card["cotista"]
            imovel_orig  = cotista_para_origem.get(cotista, "")
            energia_periodo.append({
                "cotista":       cotista,
                "imovel_uso":    imovel_uso,
                "imovel_origem": imovel_orig,
                "data":          card["data"],
                "leitura_ini":   leitura_ini,
                "leitura_fim":   card["leitura"],
                "consumo":       consumo,
                "periodo_uso":   periodo_uso,
            })

    # Build Excel
    _DARK  = "1F3864"
    _MED   = "2F5597"
    _LGRAY = "F2F2F2"
    _WHITE = "FFFFFF"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def _hdr(cell, bg=_DARK, fg=_WHITE, bold=True, sz=10, align="center", wrap=False):
        cell.font      = Font(bold=bold, color=fg, size=sz)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        return cell

    def _brl(cell, val):
        cell.value        = val if val is not None else 0.0
        cell.number_format = 'R$ #,##0.00'
        cell.alignment    = Alignment(horizontal="right", vertical="center")

    def _stripe(ws, row, ncols, bg=_LGRAY):
        if row % 2 == 0:
            for c in range(1, ncols + 1):
                ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=bg)

    def _widths(ws, ws_list):
        for i, w in enumerate(ws_list, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _merge(ws, r, c1, c2, text, **kw):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r, column=c1, value=text)
        _hdr(cell, **kw)
        return cell

    # Aba Energia
    ws = wb.create_sheet("Energia")
    _widths(ws, [24, 26, 24, 14, 14, 13, 20])
    for c, h in enumerate(["Cotista", "Imóvel (uso)", "Imóvel (cobrança)", "Med. Inicial",
                            "Med. Final", "Total (kWh)", "Período de Uso"], 1):
        _hdr(ws.cell(row=1, column=c, value=h))
    for r, row in enumerate(sorted(energia_periodo,
                                   key=lambda x: (x["imovel_origem"], x["cotista"],
                                                  x["data"] or date.min)), 2):
        ws.cell(r, 1).value = row["cotista"]
        ws.cell(r, 2).value = row["imovel_uso"]
        ws.cell(r, 3).value = row["imovel_origem"]
        ws.cell(r, 4).value = row["leitura_ini"]
        ws.cell(r, 5).value = row["leitura_fim"]
        ws.cell(r, 6).value = row["consumo"]
        ws.cell(r, 7).value = row["periodo_uso"]
        _stripe(ws, r, 7)
    ws.freeze_panes = "A2"

    # Aba Manutenção
    ws = wb.create_sheet("Manutenção")
    _widths(ws, [14, 12, 24, 42, 42, 15, 24, 24])
    for c, h in enumerate(["ID", "Data", "Imóvel", "Descrição", "Solução",
                            "Valor Total", "Custeio", "Cotista Esp."], 1):
        _hdr(ws.cell(row=1, column=c, value=h))
    for r, row in enumerate(manut_periodo, 2):
        ws.cell(r, 1).value = row["id"]
        ws.cell(r, 2).value = row["data"].strftime("%d/%m/%Y") if row["data"] else ""
        ws.cell(r, 3).value = row["imovel"]
        ws.cell(r, 4).value = row["descricao"]
        ws.cell(r, 5).value = row["solucao"]
        _brl(ws.cell(r, 6), row["valor"])
        ws.cell(r, 7).value = row["custeio"]
        ws.cell(r, 8).value = row["cotista_esp"]
        _stripe(ws, r, 8)
    ws.freeze_panes = "A2"

    # Aba Zeladoria
    ws = wb.create_sheet("Zeladoria")
    _widths(ws, [14, 12, 24, 22, 42, 15, 24, 24])
    for c, h in enumerate(["ID", "Data", "Imóvel", "Tipo", "Descrição",
                            "Valor", "Custeio", "Cotista Esp."], 1):
        _hdr(ws.cell(row=1, column=c, value=h))
    for r, row in enumerate(zel_periodo, 2):
        ws.cell(r, 1).value = row["id"]
        ws.cell(r, 2).value = row["data"].strftime("%d/%m/%Y") if row["data"] else ""
        ws.cell(r, 3).value = row["imovel"]
        ws.cell(r, 4).value = row["tipo"]
        ws.cell(r, 5).value = row["descricao"]
        _brl(ws.cell(r, 6), row["valor"])
        ws.cell(r, 7).value = row["custeio"]
        ws.cell(r, 8).value = row["cotista_esp"]
        _stripe(ws, r, 8)
    ws.freeze_panes = "A2"

    # Abas por imóvel
    imoveis_com_dados = sorted(set(
        [r["imovel"]        for r in manut_periodo] +
        [r["imovel"]        for r in zel_periodo]   +
        [r["imovel_origem"] for r in energia_periodo if r["imovel_origem"]]
    ))
    if apenas_imovel:
        imoveis_com_dados = [i for i in imoveis_com_dados if i == apenas_imovel]

    NC = 4

    for imovel in imoveis_com_dados:
        cotistas = imoveis_cotistas.get(imovel, [])
        n = max(len(cotistas), 1)

        ws = wb.create_sheet(_safe_sheet(imovel))
        _widths(ws, [28, 18, 18, 22])
        rn = 1

        _merge(ws, rn, 1, NC, imovel, bg=_DARK, sz=13)
        ws.row_dimensions[rn].height = 22
        rn += 1
        _merge(ws, rn, 1, NC,
               f"Período: {periodo_ini.strftime('%d/%m/%Y')} — {periodo_fim.strftime('%d/%m/%Y')}",
               bg=_MED, sz=10)
        rn += 2

        e_rows  = [r for r in energia_periodo if r["imovel_origem"] == imovel]
        m_all   = [r for r in manut_periodo   if r["imovel"]        == imovel]
        z_all   = [r for r in zel_periodo     if r["imovel"]        == imovel]

        m_rateio = [r for r in m_all if r["custeio"] == "Rateio entre cotistas"]
        m_cotist = [r for r in m_all if r["custeio"] == "Cotista"]
        z_rateio = [r for r in z_all if r["custeio"] == "Rateio entre cotistas"]
        z_cotist = [r for r in z_all if r["custeio"] == "Cotista"]

        tot_mr = sum(r["valor"] for r in m_rateio)
        tot_zr = sum(r["valor"] for r in z_rateio)

        headers = ["Cotista", "Energia (kWh)", "Manutenção (R$)", "Reposição de Itens (R$)"]
        for c, h in enumerate(headers, 1):
            _hdr(ws.cell(rn, c, value=h), bg=_DARK, sz=10)
        rn += 1

        linhas = cotistas if cotistas else ["(sem cotistas mapeados)"]
        total_e = total_m = total_z = 0.0

        for cot in linhas:
            e_cot = sum((r["consumo"] or 0) for r in e_rows if r["cotista"] == cot) \
                    if cotistas else sum((r["consumo"] or 0) for r in e_rows)
            m_cot = (tot_mr / n if cotistas else tot_mr) + \
                    sum(r["valor"] for r in m_cotist if r["cotista_esp"] == cot)
            z_cot = (tot_zr / n if cotistas else tot_zr) + \
                    sum(r["valor"] for r in z_cotist if r["cotista_esp"] == cot)

            total_e += e_cot
            total_m += m_cot
            total_z += z_cot

            ws.cell(rn, 1, cot).font = Font(size=10)
            ws.cell(rn, 1).alignment = Alignment(vertical="center")
            c2 = ws.cell(rn, 2)
            c2.value = round(e_cot, 1) if e_cot else 0
            c2.alignment = Alignment(horizontal="right", vertical="center")
            _brl(ws.cell(rn, 3), m_cot)
            _brl(ws.cell(rn, 4), z_cot)
            _stripe(ws, rn, NC)
            rn += 1

        for c in range(1, NC + 1):
            ws.cell(rn, c).fill = PatternFill("solid", fgColor=_MED)
        ws.cell(rn, 1, "TOTAL").font = Font(bold=True, color=_WHITE, size=10)
        ws.cell(rn, 1).fill         = PatternFill("solid", fgColor=_MED)
        ws.cell(rn, 1).alignment    = Alignment(vertical="center")
        c2 = ws.cell(rn, 2)
        c2.value     = round(total_e, 1)
        c2.font      = Font(bold=True, color=_WHITE, size=10)
        c2.fill      = PatternFill("solid", fgColor=_MED)
        c2.alignment = Alignment(horizontal="right", vertical="center")
        _brl(ws.cell(rn, 3), total_m)
        ws.cell(rn, 3).font = Font(bold=True, color=_WHITE, size=10)
        ws.cell(rn, 3).fill = PatternFill("solid", fgColor=_MED)
        _brl(ws.cell(rn, 4), total_z)
        ws.cell(rn, 4).font = Font(bold=True, color=_WHITE, size=10)
        ws.cell(rn, 4).fill = PatternFill("solid", fgColor=_MED)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ─── Flask routes ─────────────────────────────────────────────────────────────

def _cors(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    return response


@app.route("/relatorios/api/gerar")
def gerar():
    ini    = request.args.get("ini", "")
    fim    = request.args.get("fim", "")
    imovel = request.args.get("imovel") or None

    if not ini or not fim:
        return _cors(jsonify({"ok": False, "error": "Parâmetros ini e fim são obrigatórios"})), 400

    try:
        periodo_ini = date.fromisoformat(ini)
        periodo_fim = date.fromisoformat(fim)
    except ValueError as e:
        return _cors(jsonify({"ok": False, "error": f"Data inválida: {e}"})), 400

    try:
        xlsx_bytes = _gerar_relatorio(periodo_ini, periodo_fim, imovel)
        filename   = (f"Relatorio_Cotistas_{periodo_ini:%d%m}-{periodo_fim:%d%m%Y}.xlsx")
        resp = Response(
            xlsx_bytes,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
        resp.headers["Access-Control-Allow-Origin"] = "*"
        return resp
    except Exception as e:
        return _cors(jsonify({"ok": False, "error": str(e)})), 500


@app.route("/relatorios/api/gerar", methods=["OPTIONS"])
def gerar_options():
    return _cors(jsonify({}))


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5005))
    app.run(host="0.0.0.0", port=port, debug=False)
