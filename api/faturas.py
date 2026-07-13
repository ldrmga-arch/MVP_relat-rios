#!/usr/bin/env python3
"""
Faturas API — MDN Porto Rico
GET /faturas/api/gerar?ini=2026-05-27&fim=2026-06-26[&imovel=PR+RR+Q02+L14][&emissao=2026-07-01][&vencimento=2026-07-15]
Gera o Excel de faturas e retorna como download.
"""
import json, os, re, io, threading
from collections import defaultdict
from datetime import date, timedelta
import urllib.request, urllib.parse
from flask import Flask, jsonify, request, Response
from dotenv import load_dotenv

load_dotenv()

CLIENT_ID     = os.environ.get("PIPEFY_CLIENT_ID", "") or os.environ.get("CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("PIPEFY_CLIENT_SECRET", "") or os.environ.get("CLIENT_SECRET", "")
TOKEN_URL     = "https://app.pipefy.com/oauth/token"
GRAPHQL_URL   = "https://api.pipefy.com/graphql"

FASE_MANUT_OK   = "329509955"
FASE_ZEL_OK     = "329593953"
FASE_ENERGIA_OK = "340612497"
DB_COTISTAS     = "304903871"

TEMPLATE_FILE  = os.path.join(os.path.dirname(__file__), "..", "Modelo de exemplo.xlsx")
TEMPLATE_SHEET = "PR RR QD 23 LT 14"

COTISTAS_IGNORAR = {
    "Rafael França", "Marcel Sanches",
    "Gustavo Pagani", "Maria Wilma Wohlers da Silva",
}
COTISTAS_ALIAS = {
    "CARINA MUZULAN": "Carina / Clovis Muzulan",
    "Clovis Muzulan": "Carina / Clovis Muzulan",
}
TIPO_ZEL_CATEGORIA = {
    "Reposição de itens":  "REPOSIÇÃO DE ITEM",
    "Limpeza Pet":         "PRODUTOS LIMPEZA",
    "Produtos de limpeza": "PRODUTOS LIMPEZA",
}

_Q_PHASE = """
query($pid: ID!, $n: Int!, $after: String, $from: String, $to: String) {
  phase(id: $pid) {
    cards(first: $n, after: $after, filter: {created_at: {from: $from, to: $to}}) {
      pageInfo { hasNextPage endCursor }
      edges { node { id title created_at fields { name value } } }
    }
  }
}"""

_Q_PHASE_NF = """
query($pid: ID!, $n: Int!, $after: String) {
  phase(id: $pid) {
    cards(first: $n, after: $after) {
      pageInfo { hasNextPage endCursor }
      edges { node { id title created_at fields { name value } } }
    }
  }
}"""

_Q_DB = """
query($tid: ID!, $n: Int!, $after: String) {
  table_records(table_id: $tid, first: $n, after: $after) {
    pageInfo { hasNextPage endCursor }
    edges { node { title record_fields { name value array_value } } }
  }
}"""

app = Flask(__name__)


# ─── Pipefy helpers ───────────────────────────────────────────────────────────

def _get_token():
    data = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID, "client_secret": CLIENT_SECRET,
    }).encode()
    req = urllib.request.Request(TOKEN_URL, data=data, method="POST")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read())["access_token"]


def _gql(token, query, variables=None):
    body = json.dumps({"query": query, "variables": variables or {}}).encode()
    req = urllib.request.Request(GRAPHQL_URL, data=body, method="POST")
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Content-Type", "application/json")
    with urllib.request.urlopen(req, timeout=90) as r:
        d = json.loads(r.read())
    if "data" not in d:
        raise RuntimeError(json.dumps(d))
    return d["data"]


def _fetch_phase(token, phase_id, date_from=None, date_to=None):
    cards, after = [], None
    use_filter = bool(date_from and date_to)
    while True:
        try:
            if use_filter:
                d = _gql(token, _Q_PHASE, {
                    "pid": phase_id, "n": 50, "after": after,
                    "from": date_from, "to": date_to,
                })
            else:
                d = _gql(token, _Q_PHASE_NF, {"pid": phase_id, "n": 50, "after": after})
        except RuntimeError:
            if use_filter:
                use_filter, after, cards = False, None, []
                continue
            raise
        page = d["phase"]["cards"]
        for e in page["edges"]:
            cards.append(e["node"])
        if not page["pageInfo"]["hasNextPage"]:
            break
        after = page["pageInfo"]["endCursor"]
    return cards


def _fetch_db(token, table_id):
    records, after = [], None
    while True:
        d = _gql(token, _Q_DB, {"tid": table_id, "n": 50, "after": after})
        page = d["table_records"]
        for e in page["edges"]:
            records.append(e["node"])
        if not page["pageInfo"]["hasNextPage"]:
            break
        after = page["pageInfo"]["endCursor"]
    return records


# ─── Data helpers ──────────────────────────────────────────────────────────────

def _normalizar(nome):
    nome = (nome or "").strip()
    return COTISTAS_ALIAS.get(nome, nome)


def _parse_json_array(val):
    if not val:
        return ""
    try:
        lst = json.loads(val)
        return lst[0] if lst else ""
    except Exception:
        return str(val).strip('[]"\'')


def _parse_float(val):
    if not val:
        return 0.0
    nums = re.findall(r'\d+(?:[.,]\d+)?', str(val).replace(",", "."))
    return float(nums[-1].replace(",", ".")) if nums else 0.0


def _parse_energy(obs, field):
    for v in (field, obs):
        if v:
            nums = re.findall(r'\d+(?:[.,]\d+)?', str(v).strip())
            if nums:
                return _parse_float(nums[-1])
    return None


def _to_date(s):
    if not s:
        return None
    try:
        return date.fromisoformat(str(s)[:10])
    except Exception:
        return None


def _ultima_quarta_ate_dia10(year, month):
    d = date(year, month, 10)
    return d - timedelta(days=(d.weekday() - 2) % 7)


def _safe_title(name):
    for ch in "[]:*?/\\":
        name = name.replace(ch, "")
    return name[:31].strip()


# ─── Excel generation ─────────────────────────────────────────────────────────

def _gerar_excel(periodo_ini, periodo_fim, data_emissao, data_vencimento, apenas_imovel=None):
    from openpyxl import load_workbook

    # Derived periods
    _bm, _by = periodo_fim.month, periodo_fim.year
    _pm = _bm - 1 if _bm > 1 else 12
    _py = _by if _bm > 1 else _by - 1
    energia_ini = _ultima_quarta_ate_dia10(_py, _pm)
    energia_fim = _ultima_quarta_ate_dia10(_by, _bm)
    energia_periodo_str = (
        f"{energia_ini.strftime('%d/%m/%Y')} à {energia_fim.strftime('%d/%m/%Y')}"
    )
    _rm = data_emissao.month - 1 if data_emissao.month > 1 else 12
    _ry = data_emissao.year if data_emissao.month > 1 else data_emissao.year - 1
    data_ref_energia = date(_ry, _rm, 15)

    # Fetch data in parallel
    token = _get_token()
    _res = {}

    def _load(key, fn, *args, **kw):
        _res[key] = fn(token, *args, **kw)

    _energia_from = energia_ini - timedelta(days=45)
    threads = [
        threading.Thread(target=_load, args=("cot",   _fetch_db,    DB_COTISTAS)),
        threading.Thread(target=_load, args=("manut",  _fetch_phase, FASE_MANUT_OK),
                         kwargs={"date_from": periodo_ini.isoformat(), "date_to": periodo_fim.isoformat()}),
        threading.Thread(target=_load, args=("zel",    _fetch_phase, FASE_ZEL_OK),
                         kwargs={"date_from": periodo_ini.isoformat(), "date_to": periodo_fim.isoformat()}),
        threading.Thread(target=_load, args=("energ",  _fetch_phase, FASE_ENERGIA_OK),
                         kwargs={"date_from": _energia_from.isoformat(), "date_to": periodo_fim.isoformat()}),
    ]
    for t in threads: t.start()
    for t in threads: t.join()

    cotistas_raw = _res["cot"]
    manut_all    = _res["manut"]
    zel_all      = _res["zel"]
    energ_all    = _res["energ"]

    # Build cotistas map: cotas_imovel[imovel] = {cota_letter: cotista_name}
    cotas_imovel: dict = defaultdict(dict)
    cotista_para_origem: dict = {}

    for rec in cotistas_raw:
        nome_raw = (rec.get("title") or "").strip()
        if not nome_raw or nome_raw in COTISTAS_IGNORAR:
            continue
        nome = _normalizar(nome_raw)
        fmap = {f["name"]: f for f in rec.get("record_fields", [])}

        f_im = fmap.get("Imóvel")
        if not f_im:
            continue
        v = (f_im.get("value") or "").strip()
        try:
            im_list = json.loads(v) if v else []
            if not isinstance(im_list, list):
                im_list = [im_list]
        except Exception:
            im_list = [v] if v else []

        f_cota = fmap.get("Cota")
        cota = ((f_cota.get("value") or "").strip() if f_cota else "").upper()

        for imovel in im_list:
            imovel = str(imovel).strip()
            if not imovel:
                continue
            if cota:
                cotas_imovel[imovel][cota] = nome
            if nome not in cotista_para_origem:
                cotista_para_origem[nome] = imovel

    # Filter and process cards
    def in_period(d):
        return d is not None and periodo_ini <= d <= periodo_fim

    def card_fields(card):
        return {f["name"]: (f.get("value") or "") for f in card.get("fields", [])}

    manut_periodo = []
    for card in manut_all:
        dt = _to_date(card["created_at"])
        if not in_period(dt):
            continue
        f = card_fields(card)
        custeio = f.get("Responsável pelo custeio", "")
        if custeio not in ("Rateio entre cotistas", "Cotista"):
            continue
        imovel = _parse_json_array(f.get("Imóveis", ""))
        if apenas_imovel and imovel != apenas_imovel:
            continue
        manut_periodo.append({
            "imovel":      imovel,
            "valor":       _parse_float(f.get("Valor total da manutenção", "0")),
            "custeio":     custeio,
            "cotista_esp": _normalizar(_parse_json_array(f.get("Selecione o cliente", ""))),
        })

    zel_periodo = []
    for card in zel_all:
        dt = _to_date(card["created_at"])
        if not in_period(dt):
            continue
        f = card_fields(card)
        tipo = f.get("Seleção de lista", "")
        if tipo not in TIPO_ZEL_CATEGORIA:
            continue
        cobranca = f.get("Responsável pela cobrança", "") or f.get("Cobrança para", "")
        if cobranca not in ("Rateio entre cotistas", "Cotista"):
            continue
        imovel = _parse_json_array(f.get("Selecione o imóvel", ""))
        if apenas_imovel and imovel != apenas_imovel:
            continue
        zel_periodo.append({
            "imovel":      imovel,
            "tipo":        tipo,
            "categoria":   TIPO_ZEL_CATEGORIA[tipo],
            "valor":       _parse_float(f.get("Valor da cobrança", "0")),
            "custeio":     cobranca,
            "cotista_esp": _normalizar(_parse_json_array(f.get("Selecione o cotista:", ""))),
        })

    # Energy readings
    _leituras: dict = defaultdict(list)
    for card in energ_all:
        f = card_fields(card)
        imovel_uso = _parse_json_array(f.get("Selecione o imóvel", ""))
        leitura    = _parse_energy(f.get("Observação:", ""),
                                   f.get("Digite o valor no relógio de energia:", ""))
        cotista    = _normalizar(_parse_json_array(f.get("Cotista", "")))
        _leituras[imovel_uso].append({
            "data":    _to_date(card["created_at"]),
            "leitura": leitura,
            "cotista": cotista,
        })

    def in_energia_period(d):
        return d is not None and energia_ini < d <= energia_fim

    energia_periodo = []
    for imovel_uso, cards in _leituras.items():
        cards.sort(key=lambda x: x["data"] or date.min)
        for i, card in enumerate(cards):
            if not in_energia_period(card["data"]):
                continue
            prev         = cards[i - 1] if i > 0 else None
            leit_ini     = prev["leitura"] if prev else None
            consumo      = (
                card["leitura"] - leit_ini
                if card["leitura"] is not None and leit_ini is not None
                else None
            )
            cotista      = card["cotista"]
            imovel_orig  = cotista_para_origem.get(cotista, "")
            data_ini     = prev["data"] if prev else None
            periodo_uso  = (
                f"{data_ini.strftime('%d/%m')} à {card['data'].strftime('%d/%m')}"
                if data_ini and card["data"] else ""
            )
            energia_periodo.append({
                "cotista":       cotista,
                "imovel_uso":    imovel_uso,
                "imovel_origem": imovel_orig,
                "consumo":       consumo,
                "periodo":       periodo_uso,
            })

    # Excel generation
    wb = load_workbook(TEMPLATE_FILE)
    ws_tpl = wb[TEMPLATE_SHEET]

    imoveis = ([apenas_imovel] if apenas_imovel else sorted(set(
        [r["imovel"]        for r in manut_periodo] +
        [r["imovel"]        for r in zel_periodo]   +
        [r["imovel_origem"] for r in energia_periodo if r["imovel_origem"]]
    )))

    for imovel in imoveis:
        cotas = cotas_imovel.get(imovel, {})
        n = max(len(cotas), 1)
        cot2cota = {v: k for k, v in cotas.items()}

        # Manutenção
        m_all    = [r for r in manut_periodo if r["imovel"] == imovel]
        m_rateio = sum(r["valor"] for r in m_all if r["custeio"] == "Rateio entre cotistas")
        m_ind    = [r for r in m_all if r["custeio"] == "Cotista"]

        serv_mdn: dict = {}
        for cota, cotista in cotas.items():
            ind = sum(r["valor"] for r in m_ind if r["cotista_esp"] == cotista)
            serv_mdn[cota] = round(m_rateio / n + ind, 2)

        # Zeladoria
        repos_item:   dict = {}
        prod_limpeza: dict = {}

        for categoria, store in [("REPOSIÇÃO DE ITEM", repos_item),
                                  ("PRODUTOS LIMPEZA",  prod_limpeza)]:
            z_all = [r for r in zel_periodo
                     if r["imovel"] == imovel and r["categoria"] == categoria]
            z_rat = sum(r["valor"] for r in z_all
                        if r["custeio"] == "Rateio entre cotistas")
            for cota, cotista in cotas.items():
                z_ind = sum(r["valor"] for r in z_all
                            if r["custeio"] == "Cotista"
                            and r["cotista_esp"] == cotista)
                store[cota] = round(z_rat / n + z_ind, 2)

        # Energia
        e_por_cota: dict = defaultdict(float)
        e_periodos: dict  = defaultdict(list)
        for r in energia_periodo:
            if r["imovel_origem"] != imovel:
                continue
            cota = cot2cota.get(r["cotista"], "")
            if cota and r["consumo"] is not None:
                e_por_cota[cota] += r["consumo"]
                e_periodos[cota].append(r["periodo"])

        # Copy template sheet
        ws = wb.copy_worksheet(ws_tpl)
        ws.title = _safe_title(imovel)

        sorted_cotas = sorted(cotas.keys())

        ws["A8"] = f"MONDONEX - {imovel}"
        ws["B8"] = n
        ws["D5"] = data_emissao
        ws["D8"] = data_vencimento

        itens = [
            (serv_mdn,    19),
            (repos_item,  22),
            (prod_limpeza, 25),
        ]
        for data_dict, base_row in itens:
            for i, cota in enumerate(sorted_cotas):
                r = base_row + i
                ws.cell(r, 2).value = cota
                ws.cell(r, 3).value = data_dict.get(cota, 0)

        for data_dict, base_row in itens:
            for i in range(len(sorted_cotas), 3):
                ws.cell(base_row + i, 2).value = None
                ws.cell(base_row + i, 3).value = None

        # Energy (CI table)
        for r in range(2, 30):
            for c in range(6, 13):
                ws.cell(r, c).value = None

        r_ci = 2
        for cota in sorted_cotas:
            consumo = e_por_cota.get(cota, 0.0)
            if consumo <= 0:
                continue
            periodos = e_periodos.get(cota, [])
            if periodos:
                periodo_str = periodos[0] if len(periodos) == 1 else \
                              f"{periodos[0].split(' à ')[0]} à {periodos[-1].split(' à ')[-1]}"
            else:
                periodo_str = energia_periodo_str

            ws.cell(r_ci, 6).value  = cota
            ws.cell(r_ci, 8).value  = data_ref_energia
            ws.cell(r_ci, 9).value  = round(consumo, 2)
            ws.cell(r_ci, 10).value = "1.4.12 - Energia - CI"
            ws.cell(r_ci, 11).value = energia_periodo_str
            r_ci += 1

    # Remove template sheets
    for sheet_name in ["PR RR QD02 LT14", TEMPLATE_SHEET]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ─── Flask routes ─────────────────────────────────────────────────────────────

def _cors(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    return response


@app.route("/faturas/api/gerar")
def gerar():
    ini        = request.args.get("ini", "")
    fim        = request.args.get("fim", "")
    imovel     = request.args.get("imovel") or None
    emissao    = request.args.get("emissao") or None
    vencimento = request.args.get("vencimento") or None

    if not ini or not fim:
        return _cors(jsonify({"ok": False, "error": "Parâmetros ini e fim são obrigatórios"})), 400

    try:
        periodo_ini    = date.fromisoformat(ini)
        periodo_fim    = date.fromisoformat(fim)
        data_emissao   = date.fromisoformat(emissao)    if emissao    else date.today()
        data_vencimento= date.fromisoformat(vencimento) if vencimento else date.today().replace(day=15)
    except ValueError as e:
        return _cors(jsonify({"ok": False, "error": f"Data inválida: {e}"})), 400

    try:
        xlsx_bytes = _gerar_excel(periodo_ini, periodo_fim, data_emissao, data_vencimento, imovel)
        filename   = f"Faturas_{periodo_ini:%d%m}-{periodo_fim:%d%m%Y}.xlsx"
        resp = Response(
            xlsx_bytes,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
        resp.headers["Access-Control-Allow-Origin"] = "*"
        return resp
    except Exception as e:
        return _cors(jsonify({"ok": False, "error": str(e)})), 500


@app.route("/faturas/api/gerar", methods=["OPTIONS"])
def gerar_options():
    return _cors(jsonify({}))


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5004))
    app.run(host="0.0.0.0", port=port, debug=False)
